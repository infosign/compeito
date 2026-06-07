"""XLSX Export Service — exports a CFDocument to the OpenSALT Excel format.

OpenSALT's only full-fidelity table interchange format is a 3-sheet Excel
workbook (CF Doc / CF Item / CF Association). This service produces a workbook
that OpenSALT's Excel importer (``/salt/excel/import``) consumes, and that the
companion :mod:`xlsx_import_service` reads back. See
docs/dev/round_trip_status.md (OpenSALT section) for the format reference.

Column layouts mirror OpenSALT's ``ExcelExport.php`` / ``ExcelImport.php``:

- **CF Doc** (A-P): identifier, creator, title, lastChangeDate, officialSourceURL,
  publisher, description, subject (``|``-joined), language, version, adoptionStatus,
  statusStartDate, statusEndDate, licenseTitle, licenseText, notes
- **CF Item** (A-L): identifier, fullStatement, humanCodingScheme, smartLevel,
  listEnumeration, abbreviatedStatement, conceptKeywords, notes, language,
  educationLevel, CFItemType, license
- **CF Association** (A-J): identifier, originNodeURI, originNodeIdentifier,
  originNodeHumanCodingScheme, associationType, destinationNodeURI,
  destinationNodeIdentifier, destinationNodeHumanCodingScheme,
  associationGroupIdentifier, associationGroupName

Known gaps vs. CASE (documented in round_trip_status.md): compeito has no
``notes`` column on CFItem/CFDocument, so those cells are emitted empty. The
hierarchy is expressed via ``smartLevel`` (so isChildOf associations are NOT
repeated in the CF Association sheet); ``smartLevel`` segments are 1-based
sibling positions.
"""

from __future__ import annotations

import io
import uuid
from collections import defaultdict

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from src.models.cf_association import CFAssociation
from src.services.csv_export_service import _jsonb_list_to_csv, _load_document_tree

CF_DOC_HEADER = [
    "identifier",
    "creator",
    "title",
    "lastChangeDate",
    "officialSourceURL",
    "publisher",
    "description",
    "subject",
    "language",
    "version",
    "adoptionStatus",
    "statusStartDate",
    "statusEndDate",
    "licenseTitle",
    "licenseText",
    "notes",
]

CF_ITEM_HEADER = [
    "identifier",
    "fullStatement",
    "humanCodingScheme",
    "smartLevel",
    "listEnumeration",
    "abbreviatedStatement",
    "conceptKeywords",
    "notes",
    "language",
    "educationLevel",
    "CFItemType",
    "license",
]

CF_ASSOCIATION_HEADER = [
    "identifier",
    "originNodeURI",
    "originNodeIdentifier",
    "originNodeHumanCodingScheme",
    "associationType",
    "destinationNodeURI",
    "destinationNodeIdentifier",
    "destinationNodeHumanCodingScheme",
    "associationGroupIdentifier",
    "associationGroupName",
]


def _subject_to_pipe(subject) -> str:
    """CFDocument.subject (JSONB list) -> pipe-joined string (OpenSALT style)."""
    if isinstance(subject, list):
        return "|".join(str(s) for s in subject if str(s).strip())
    return ""


def _compute_smart_levels(
    ordered: list[tuple[object, str | None, int | None]],
) -> dict[str, str]:
    """Assign an OpenSALT ``smartLevel`` to each item.

    ``ordered`` is the depth-first pre-order list from ``_load_document_tree``,
    so a parent always precedes its children. Each segment is a 1-based sibling
    position (``1``, ``1.1``, ``1.2``, ``2`` ...).
    """
    smart: dict[str, str] = {}
    counter: dict[str, int] = defaultdict(int)
    for item, parent_ident, _seq in ordered:
        ident = str(item.identifier)
        parent_sl = smart.get(parent_ident, "") if parent_ident else ""
        counter[parent_sl] += 1
        idx = counter[parent_sl]
        smart[ident] = f"{parent_sl}.{idx}" if parent_sl else str(idx)
    return smart


async def export_xlsx(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    doc_identifier: uuid.UUID,
) -> bytes:
    """Export a document to an OpenSALT-format Excel workbook.

    Returns:
        The ``.xlsx`` file as bytes.

    Raises:
        ValueError: If the document is not found.
    """
    doc, ordered = await _load_document_tree(session, tenant_id, doc_identifier)
    smart_levels = _compute_smart_levels(ordered)

    # Non-isChildOf associations go in the CF Association sheet (hierarchy is
    # carried by smartLevel in the CF Item sheet).
    result = await session.execute(
        select(CFAssociation)
        .options(joinedload(CFAssociation.association_grouping))
        .where(
            CFAssociation.cf_document_id == doc.id,
            CFAssociation.association_type != "isChildOf",
        )
    )
    assocs = list(result.scalars().unique().all())

    wb = Workbook()
    # Replace the default sheet with our three named sheets.
    wb.remove(wb.active)

    # --- CF Doc ---
    ws_doc = wb.create_sheet("CF Doc")
    ws_doc.append(CF_DOC_HEADER)
    lic = doc.license
    ws_doc.append(
        [
            str(doc.identifier),
            doc.creator or "",
            doc.title or "",
            doc.last_change_date_time.isoformat() if doc.last_change_date_time else "",
            doc.official_source_url or "",
            doc.publisher or "",
            doc.description or "",
            _subject_to_pipe(doc.subject),
            doc.language or "",
            doc.version or "",
            doc.adoption_status or "",
            str(doc.status_start_date) if doc.status_start_date else "",
            str(doc.status_end_date) if doc.status_end_date else "",
            lic.title if lic else "",
            (lic.license_text or "") if lic else "",
            "",  # notes — not stored on CFDocument
        ]
    )

    # --- CF Item ---
    ws_item = wb.create_sheet("CF Item")
    ws_item.append(CF_ITEM_HEADER)
    for item, _parent_ident, _seq in ordered:
        ws_item.append(
            [
                str(item.identifier),
                item.full_statement or "",
                item.human_coding_scheme or "",
                smart_levels.get(str(item.identifier), ""),
                item.list_enumeration or "",
                item.abbreviated_statement or "",
                _jsonb_list_to_csv(item.concept_keywords),
                "",  # notes — not stored on CFItem
                item.language or "",
                _jsonb_list_to_csv(item.education_level),
                item.item_type.title if item.item_type else "",
                "",  # license — managed at the document level
            ]
        )

    # --- CF Association ---
    ws_assoc = wb.create_sheet("CF Association")
    ws_assoc.append(CF_ASSOCIATION_HEADER)
    for a in assocs:
        grouping = a.association_grouping
        ws_assoc.append(
            [
                str(a.identifier),
                a.origin_node_uri or "",
                a.origin_node_identifier or "",
                "",  # originNodeHumanCodingScheme — not stored
                a.association_type or "",
                a.destination_node_uri or "",
                a.destination_node_identifier or "",
                "",  # destinationNodeHumanCodingScheme — not stored
                str(grouping.identifier) if grouping else "",
                grouping.title if grouping else "",
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
