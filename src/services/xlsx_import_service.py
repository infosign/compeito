"""XLSX Import Service — imports an OpenSALT-format Excel workbook.

Reads the 3-sheet OpenSALT workbook (CF Doc / CF Item / CF Association; see
:mod:`xlsx_export_service` and docs/dev/round_trip_status.md for the layout) and
persists it.

Strategy: the CF Doc + CF Item sheets are converted in-memory to compeito's
**custom CSV** format and fed through the existing, well-tested
:func:`csv_import_service.import_csv` path — this reuses all item upsert,
CFItemType find-or-create, education-level, and isChildOf hierarchy logic. The
hierarchy carried by ``smartLevel`` is translated into the custom format's
``parentIdentifier`` + ``sequenceNumber`` columns. The CF Association sheet's
non-isChildOf associations are then imported in a dedicated second pass
(isChildOf is already materialised from smartLevel, so it is skipped here).

Known gaps (documented in round_trip_status.md): compeito has no ``notes``
column on CFItem/CFDocument, so the ``notes`` cells are ignored.
"""

from __future__ import annotations

import csv
import io
import uuid
from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.cf_association import CFAssociation
from src.models.cf_association_grouping import CFAssociationGrouping
from src.models.cf_document import CFDocument
from src.services.csv_export_service import HEADER as CUSTOM_HEADER
from src.services.csv_import_service import (
    ImportReport,
    _build_uri,
    _is_valid_uuid,
    _now_utc,
    import_csv,
)

# CF Doc sheet column index → metadata key for the custom CSV "#" header rows.
# (Column order mirrors xlsx_export_service.CF_DOC_HEADER.)
_DOC_META_BY_COL = {
    0: "identifier",
    1: "creator",
    2: "title",
    # 3: lastChangeDate — re-stamped on import, ignored
    4: "official_source_url",
    5: "publisher",
    6: "description",
    # 7: subject — handled separately (pipe-split → multi-value row)
    8: "language",
    9: "version",
    10: "adoption_status",
    11: "status_start_date",
    12: "status_end_date",
    13: "license",  # licenseTitle
    # 14: licenseText — not representable in custom #license row
    # 15: notes — not stored on CFDocument
}

# CF Item sheet column indices (mirrors xlsx_export_service.CF_ITEM_HEADER).
_ITEM_IDENTIFIER = 0
_ITEM_FULL_STATEMENT = 1
_ITEM_HCS = 2
_ITEM_SMART_LEVEL = 3
_ITEM_LIST_ENUM = 4
_ITEM_ABBREV = 5
_ITEM_CONCEPT_KEYWORDS = 6
# 7: notes — not stored
_ITEM_LANGUAGE = 8
_ITEM_EDUCATION_LEVEL = 9
_ITEM_CF_ITEM_TYPE = 10
# 11: license — managed at the document level


def _cell(value) -> str:
    """openpyxl cell value → trimmed string ('' for None)."""
    if value is None:
        return ""
    return str(value).strip()


def _row_values(ws, max_col: int) -> list[list[str]]:
    """All rows of a worksheet as lists of strings, padded to max_col."""
    rows: list[list[str]] = []
    for raw in ws.iter_rows(values_only=True):
        row = [_cell(c) for c in raw]
        if len(row) < max_col:
            row += [""] * (max_col - len(row))
        rows.append(row)
    return rows


def _smartlevel_parent(sl: str) -> str:
    return sl.rsplit(".", 1)[0] if "." in sl else ""


def _smartlevel_seq(sl: str) -> str:
    last = sl.rsplit(".", 1)[-1]
    return last if last.isdigit() else ""


def _build_custom_csv(doc_row: list[str], item_rows: list[list[str]]) -> bytes:
    """Convert CF Doc + CF Item sheet rows to compeito custom-format CSV bytes."""
    out = io.StringIO()
    writer = csv.writer(out, lineterminator="\n")

    # --- metadata "#" rows from CF Doc ---
    for col, key in sorted(_DOC_META_BY_COL.items()):
        val = doc_row[col] if col < len(doc_row) else ""
        if val:
            writer.writerow([f"#{key}", val])
    subject = doc_row[7] if len(doc_row) > 7 else ""
    if subject:
        writer.writerow(["#subject"] + [s.strip() for s in subject.split("|") if s.strip()])

    # --- ensure every item has an identifier, then map smartLevel → identifier ---
    smart_to_ident: dict[str, str] = {}
    for r in item_rows:
        if not r[_ITEM_IDENTIFIER] or not _is_valid_uuid(r[_ITEM_IDENTIFIER]):
            r[_ITEM_IDENTIFIER] = str(uuid.uuid4())
        sl = r[_ITEM_SMART_LEVEL]
        if sl:
            smart_to_ident[sl] = r[_ITEM_IDENTIFIER]

    # --- header + data rows in custom HEADER order ---
    writer.writerow(CUSTOM_HEADER)
    for r in item_rows:
        sl = r[_ITEM_SMART_LEVEL]
        parent_sl = _smartlevel_parent(sl)
        writer.writerow(
            [
                r[_ITEM_IDENTIFIER],  # Identifier
                r[_ITEM_FULL_STATEMENT],  # fullStatement
                r[_ITEM_HCS],  # humanCodingScheme
                smart_to_ident.get(parent_sl, ""),  # parentIdentifier
                _smartlevel_seq(sl),  # sequenceNumber
                r[_ITEM_CF_ITEM_TYPE],  # CFItemType
                r[_ITEM_EDUCATION_LEVEL],  # educationLevel
                r[_ITEM_CONCEPT_KEYWORDS],  # conceptKeywords
                r[_ITEM_ABBREV],  # abbreviatedStatement
                r[_ITEM_LANGUAGE],  # language
                r[_ITEM_LIST_ENUM],  # listEnumeration
                "",  # license (document-level)
                "",  # statusStartDate (item-level; not in OpenSALT item sheet)
                "",  # statusEndDate
            ]
        )

    return out.getvalue().encode("utf-8")


async def _import_associations(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    doc: CFDocument,
    assoc_rows: list[list[str]],
    now: datetime,
    report: ImportReport,
) -> None:
    """Import non-isChildOf associations from the CF Association sheet rows.

    assoc_rows excludes the header. Columns mirror
    xlsx_export_service.CF_ASSOCIATION_HEADER.
    """
    grouping_cache: dict[str, CFAssociationGrouping] = {}

    for r in assoc_rows:
        ident_raw, origin_uri, origin_id, _origin_hcs, atype, dest_uri, dest_id, _dest_hcs, group_id, group_name = r[
            :10
        ]

        if not atype or atype == "isChildOf":
            continue  # hierarchy is materialised from smartLevel
        if not origin_id or not dest_id:
            report.warnings.append(f"Association skipped (missing origin/destination): type '{atype}'")
            continue

        ident = uuid.UUID(ident_raw) if _is_valid_uuid(ident_raw) else uuid.uuid4()

        # Skip if this association already exists in the document.
        existing = await session.execute(
            select(CFAssociation).where(
                CFAssociation.cf_document_id == doc.id,
                CFAssociation.identifier == ident,
            )
        )
        if existing.scalar_one_or_none() is not None:
            continue

        # Resolve a grouping (find-or-create by identifier within the tenant —
        # CFAssociationGrouping is a tenant-wide lookup, not document-scoped).
        grouping: CFAssociationGrouping | None = None
        if group_id and _is_valid_uuid(group_id):
            gkey = group_id
            grouping = grouping_cache.get(gkey)
            if grouping is None:
                gident = uuid.UUID(group_id)
                gres = await session.execute(
                    select(CFAssociationGrouping).where(
                        CFAssociationGrouping.tenant_id == tenant_id,
                        CFAssociationGrouping.identifier == gident,
                    )
                )
                grouping = gres.scalar_one_or_none()
                if grouping is None:
                    grouping = CFAssociationGrouping(
                        id=uuid.uuid4(),
                        tenant_id=tenant_id,
                        identifier=gident,
                        uri=_build_uri(tenant_id, gident),
                        title=group_name or str(gident),
                        last_change_date_time=now,
                    )
                    session.add(grouping)
                grouping_cache[gkey] = grouping

        def _uri(raw_uri: str, node_id: str) -> str:
            if raw_uri:
                return raw_uri
            if _is_valid_uuid(node_id):
                return _build_uri(tenant_id, uuid.UUID(node_id))
            return node_id

        assoc = CFAssociation(
            id=uuid.uuid4(),
            tenant_id=tenant_id,
            cf_document_id=doc.id,
            identifier=ident,
            uri=_build_uri(tenant_id, ident),
            association_type=atype,
            origin_node_uri=_uri(origin_uri, origin_id),
            origin_node_identifier=origin_id,
            destination_node_uri=_uri(dest_uri, dest_id),
            destination_node_identifier=dest_id,
            cf_association_grouping_id=grouping.id if grouping else None,
            last_change_date_time=now,
        )
        session.add(assoc)
        report.associations_created += 1


async def import_xlsx(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    xlsx_data: bytes,
    *,
    doc_identifier: uuid.UUID | None = None,
) -> ImportReport:
    """Import an OpenSALT-format Excel workbook into the database.

    Args:
        session: Async database session (caller manages the transaction).
        tenant_id: Target tenant UUID.
        xlsx_data: Raw ``.xlsx`` file bytes.
        doc_identifier: Optional existing CFDocument UUID (update in place).

    Returns:
        ImportReport with counts and warnings.

    Raises:
        ValueError: If the workbook is malformed or required sheets are missing.
    """
    try:
        wb = load_workbook(io.BytesIO(xlsx_data), read_only=True, data_only=True)
    except Exception as e:  # openpyxl raises various errors for non-xlsx input
        raise ValueError(f"Could not read .xlsx workbook: {e}") from e

    for name in ("CF Doc", "CF Item"):
        if name not in wb.sheetnames:
            raise ValueError(f"Workbook is missing the required '{name}' sheet")

    doc_rows = _row_values(wb["CF Doc"], 16)
    if len(doc_rows) < 2:
        raise ValueError("'CF Doc' sheet has no data row")
    doc_row = doc_rows[1]  # row 1 is the header

    item_rows_all = _row_values(wb["CF Item"], 12)
    # Drop header row; keep rows that have a fullStatement.
    item_rows = [r for r in item_rows_all[1:] if r[_ITEM_FULL_STATEMENT]]

    assoc_rows: list[list[str]] = []
    if "CF Association" in wb.sheetnames:
        assoc_all = _row_values(wb["CF Association"], 10)
        assoc_rows = assoc_all[1:]

    wb.close()

    # --- items + hierarchy + doc via the custom-CSV import path ---
    csv_bytes = _build_custom_csv(doc_row, item_rows)
    report = await import_csv(session, tenant_id, csv_bytes, doc_identifier=doc_identifier, profile="custom")
    await session.flush()

    # --- non-isChildOf associations ---
    if assoc_rows:
        res = await session.execute(
            select(CFDocument).where(
                CFDocument.tenant_id == tenant_id,
                CFDocument.identifier == uuid.UUID(report.document_identifier),
            )
        )
        doc = res.scalar_one()
        await _import_associations(session, tenant_id, doc, assoc_rows, _now_utc(), report)

    return report
