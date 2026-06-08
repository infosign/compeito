"""Tests for the CASE v1.1 notes / alternativeLabel / extensions fields.

notes (CFItem/CFDocument/CFAssociation), alternativeLabel (CFItem) and
extensions (all entities) were added to close gaps against the CASE v1.1
information model. These tests verify they survive CASE-JSON import → API
serialize, the custom CSV round-trip, and the OpenSALT xlsx round-trip (notes
only — alternativeLabel/extensions have no column in the OpenSALT formats).
"""

from __future__ import annotations

import io
import uuid

from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.tenant import Tenant
from src.services import case_query_service, cf_view_service
from src.services.case_import_service import import_case_from_dict
from src.services.csv_export_service import export_csv
from src.services.csv_import_service import import_csv
from src.services.xlsx_export_service import export_xlsx
from src.services.xlsx_import_service import import_xlsx

TENANT_ID = uuid.UUID("11111111-1111-1111-1111-111111111111")
DOC = "dddddddd-0000-0000-0000-000000000001"
ITEM = "10000000-0000-0000-0000-000000000001"
ITEM2 = "10000000-0000-0000-0000-000000000002"
ASSOC = "20000000-0000-0000-0000-000000000001"
ITYPE = "30000000-0000-0000-0000-000000000001"


def _pkg() -> dict:
    return {
        "CFPackage": {
            "extensions": {"pkgExt": "p"},
            "CFDocument": {
                "identifier": DOC,
                "uri": f"https://example.com/uri/{DOC}",
                "title": "Notes Doc",
                "creator": "T",
                "notes": "doc level notes",
                "extensions": {"docExt": "dv"},
                "lastChangeDateTime": "2025-01-01T00:00:00Z",
            },
            "CFItems": [
                {
                    "identifier": ITEM,
                    "uri": f"https://example.com/uri/{ITEM}",
                    "fullStatement": "Item one",
                    "alternativeLabel": "alt one",
                    "notes": "item notes",
                    "extensions": {"itemExt": 1},
                    "CFItemType": "Skill",
                    "CFItemTypeURI": {"identifier": ITYPE, "uri": f"https://example.com/uri/{ITYPE}", "title": "Skill"},
                    "lastChangeDateTime": "2025-01-01T00:00:00Z",
                },
                {
                    "identifier": ITEM2,
                    "uri": f"https://example.com/uri/{ITEM2}",
                    "fullStatement": "Item two",
                    "lastChangeDateTime": "2025-01-01T00:00:00Z",
                },
            ],
            "CFAssociations": [
                {
                    "identifier": ASSOC,
                    "uri": f"https://example.com/uri/{ASSOC}",
                    "associationType": "isRelatedTo",
                    "originNodeURI": {"identifier": ITEM, "uri": "u", "title": "o"},
                    "destinationNodeURI": {"identifier": ITEM2, "uri": "u", "title": "d"},
                    "notes": "assoc notes",
                    "extensions": {"assocExt": True},
                    "lastChangeDateTime": "2025-01-01T00:00:00Z",
                }
            ],
            "CFDefinitions": {
                "extensions": {"defExt": "d"},
                "CFItemTypes": [
                    {
                        "identifier": ITYPE,
                        "uri": f"https://example.com/uri/{ITYPE}",
                        "title": "Skill",
                        "hierarchyCode": "1",
                        "extensions": {"typeExt": "x"},
                        "lastChangeDateTime": "2025-01-01T00:00:00Z",
                    }
                ],
            },
        }
    }


class TestCaseJsonRoundTrip:
    async def test_notes_altlabel_extensions_persist_and_serialize(self, db_session: AsyncSession):
        db_session.add(Tenant(id=TENANT_ID, name="T", is_private=False))
        await db_session.flush()
        await import_case_from_dict(db_session, TENANT_ID, _pkg())
        await db_session.flush()

        # Standalone API serializers
        item = await case_query_service.get_cf_item(db_session, TENANT_ID, uuid.UUID(ITEM))
        d = item.model_dump(by_alias=True)
        assert d["notes"] == "item notes"
        assert d["alternativeLabel"] == "alt one"
        assert d["extensions"] == {"itemExt": 1}

        doc = await case_query_service.get_cf_document(db_session, TENANT_ID, uuid.UUID(DOC))
        dd = doc.model_dump(by_alias=True)
        assert dd["notes"] == "doc level notes"
        assert dd["extensions"] == {"docExt": "dv"}

        assoc = await case_query_service.get_cf_association(db_session, TENANT_ID, uuid.UUID(ASSOC))
        da = assoc.model_dump(by_alias=True)
        assert da["notes"] == "assoc notes"
        assert da["extensions"] == {"assocExt": True}

        # Package serializer (CFPckg variants + CFDefinitions lookups)
        pkg = await cf_view_service.get_cf_package(db_session, TENANT_ID, uuid.UUID(DOC))
        pd = pkg.model_dump(by_alias=True)
        pitem = next(i for i in pd["CFItems"] if i["identifier"] == ITEM)
        assert pitem["notes"] == "item notes" and pitem["alternativeLabel"] == "alt one"
        assert pitem["extensions"] == {"itemExt": 1}
        itype = pd["CFDefinitions"]["CFItemTypes"][0]
        assert itype["extensions"] == {"typeExt": "x"}
        # Container-level extensions (CFPackage / CFDefinitions).
        assert pd["extensions"] == {"pkgExt": "p"}
        assert pd["CFDefinitions"]["extensions"] == {"defExt": "d"}


class TestCustomCsvRoundTrip:
    async def test_notes_and_alt_label_survive_csv(self, db_session: AsyncSession):
        db_session.add(Tenant(id=TENANT_ID, name="T", is_private=False))
        await db_session.flush()
        await import_case_from_dict(db_session, TENANT_ID, _pkg())
        await db_session.flush()

        csv_str = await export_csv(db_session, TENANT_ID, uuid.UUID(DOC))
        # Header carries the new columns; #notes metadata present.
        assert "alternativeLabel" in csv_str and "notes" in csv_str
        assert "#notes,doc level notes" in csv_str
        assert "item notes" in csv_str and "alt one" in csv_str

        # Re-import the exported custom CSV into a fresh tenant; fields persist.
        t2 = uuid.UUID("22222222-2222-2222-2222-222222222222")
        db_session.add(Tenant(id=t2, name="T2", is_private=False))
        await db_session.flush()
        await import_csv(db_session, t2, csv_str.encode("utf-8"))
        await db_session.flush()
        item = await case_query_service.get_cf_item(db_session, t2, uuid.UUID(ITEM))
        d = item.model_dump(by_alias=True)
        assert d["notes"] == "item notes"
        assert d["alternativeLabel"] == "alt one"


class TestXlsxRoundTripNotes:
    async def test_item_and_doc_notes_survive_xlsx(self, db_session: AsyncSession):
        db_session.add(Tenant(id=TENANT_ID, name="T", is_private=False))
        await db_session.flush()
        await import_case_from_dict(db_session, TENANT_ID, _pkg())
        await db_session.flush()

        data = await export_xlsx(db_session, TENANT_ID, uuid.UUID(DOC))
        wb = load_workbook(io.BytesIO(data))
        # CF Item notes in column H (index 8, 1-based), CF Doc notes in column P (16).
        item_rows = list(wb["CF Item"].iter_rows(min_row=2, values_only=True))
        notes_by_stmt = {r[1]: r[7] for r in item_rows}
        assert notes_by_stmt["Item one"] == "item notes"
        assert wb["CF Doc"].cell(2, 16).value == "doc level notes"

        # Re-import the workbook into a fresh tenant; notes persist.
        t2 = uuid.UUID("33333333-3333-3333-3333-333333333333")
        db_session.add(Tenant(id=t2, name="T3", is_private=False))
        await db_session.flush()
        await import_xlsx(db_session, t2, data)
        await db_session.flush()
        item = await case_query_service.get_cf_item(db_session, t2, uuid.UUID(ITEM))
        assert item.model_dump(by_alias=True)["notes"] == "item notes"
        doc = await case_query_service.get_cf_document(db_session, t2, uuid.UUID(DOC))
        assert doc.model_dump(by_alias=True)["notes"] == "doc level notes"


class TestStrictPackageMode:
    """GET /CFPackages/{id}?strict=1 omits the package-context-only properties
    (CFDocument.CFPackageURI, CFItems[].CFDocumentURI) for official-schema
    conformance; the default response keeps them (OpenCASE/OpenSALT interop)."""

    async def test_strict_omits_package_context_uris(self, db_session: AsyncSession, db_client):
        db_session.add(Tenant(id=TENANT_ID, name="T", is_private=False))
        await db_session.flush()
        await import_case_from_dict(db_session, TENANT_ID, _pkg())
        await db_session.flush()

        # Default: package-context URIs present.
        r = await db_client.get(f"/{TENANT_ID}/ims/case/v1p1/CFPackages/{DOC}")
        assert r.status_code == 200
        body = r.json()
        assert "CFPackageURI" in body["CFDocument"]
        assert all("CFDocumentURI" in it for it in body["CFItems"])

        # Strict: those keys are stripped.
        rs = await db_client.get(f"/{TENANT_ID}/ims/case/v1p1/CFPackages/{DOC}?strict=1")
        assert rs.status_code == 200
        sb = rs.json()
        assert "CFPackageURI" not in sb["CFDocument"]
        assert all("CFDocumentURI" not in it for it in sb["CFItems"])
        # Real fields still present.
        assert sb["CFDocument"]["title"] == "Notes Doc"
        assert sb["extensions"] == {"pkgExt": "p"}
