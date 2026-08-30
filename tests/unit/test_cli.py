"""Tests for CLI commands (Issue #39)."""

import asyncio
import uuid
from datetime import datetime, timezone

import pytest
from click.testing import CliRunner
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.pool import NullPool

from src.config import settings
from src.models.cf_association import CFAssociation
from src.models.cf_document import CFDocument
from src.models.tenant import Tenant

NOW = datetime(2025, 1, 1, tzinfo=timezone.utc)
TENANT_ID = uuid.UUID("11111111-1111-1111-1111-111111111111")
DOC_IDENT = uuid.UUID("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa")

_CLEANUP_TABLES = [
    "cf_rubric_criterion_levels",
    "cf_rubric_criteria",
    "cf_rubrics",
    "cf_associations",
    "cf_items",
    "cf_documents",
    "cf_association_groupings",
    "cf_concepts",
    "cf_subjects",
    "cf_item_types",
    "cf_licenses",
    "tenants",
]


# ---------------------------------------------------------------------------
# Sync DB helpers (CLI tests need sync fixtures because CLI uses asyncio.run)
# ---------------------------------------------------------------------------


async def _db_exec(callback):
    """Run an async callback with a committed DB session."""
    engine = create_async_engine(settings.database_url, poolclass=NullPool)
    async with engine.connect() as conn:
        session = AsyncSession(bind=conn, expire_on_commit=False)
        try:
            await callback(session)
            await session.commit()
        finally:
            await session.close()
    await engine.dispose()


async def _db_cleanup():
    """Delete all test data from DB."""
    engine = create_async_engine(settings.database_url, poolclass=NullPool)
    async with engine.connect() as conn:
        for table in _CLEANUP_TABLES:
            await conn.execute(text(f"DELETE FROM {table}"))
        await conn.commit()
    await engine.dispose()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _english_cli_locale(monkeypatch):
    """Force English CLI output regardless of host LANG.

    Without this, hosts with LANG=ja_JP.UTF-8 (e.g., macOS default) make the CLI
    emit Japanese messages and break tests that assert on English strings.
    """
    monkeypatch.setenv("LANG", "C")


@pytest.fixture
def runner():
    return CliRunner()


@pytest.fixture
def env_docker(monkeypatch):
    """Set DATABASE_URL environment variable for Docker mode."""
    monkeypatch.setenv("DATABASE_URL", settings.database_url)


@pytest.fixture
def clean_db():
    """Ensure DB is clean before and after test."""
    asyncio.run(_db_cleanup())
    yield
    asyncio.run(_db_cleanup())


@pytest.fixture
def test_tenant(clean_db):
    """Create a test tenant committed to DB."""

    async def _create(session):
        session.add(Tenant(id=TENANT_ID, name="Test Tenant", is_private=False))

    asyncio.run(_db_exec(_create))
    return TENANT_ID


@pytest.fixture
def test_document(test_tenant):
    """Create a test document committed to DB."""

    async def _create(session):
        session.add(
            CFDocument(
                tenant_id=TENANT_ID,
                identifier=DOC_IDENT,
                uri=f"https://example.com/uri/{DOC_IDENT}",
                title="Test Document",
                creator="Test Creator",
                language="ja",
                last_change_date_time=NOW,
            )
        )

    asyncio.run(_db_exec(_create))
    return DOC_IDENT


ASSOC_IDENT = uuid.UUID("dddddddd-dddd-dddd-dddd-dddddddddddd")


@pytest.fixture
def test_association(test_document):
    """A replacedBy association, committed to DB."""

    async def _create(session):
        session.add(
            CFAssociation(
                tenant_id=TENANT_ID,
                cf_document_id=(await session.execute(select(CFDocument).where(CFDocument.identifier == DOC_IDENT)))
                .scalar_one()
                .id,
                identifier=ASSOC_IDENT,
                uri=f"https://example.com/assoc/{ASSOC_IDENT}",
                association_type="replacedBy",
                origin_node_uri="https://example.com/uri/old",
                origin_node_identifier="11111111-2222-3333-4444-555555555555",
                origin_node_title="Old item",
                destination_node_uri="https://example.com/uri/new",
                destination_node_identifier="66666666-7777-8888-9999-000000000000",
                destination_node_title="New item",
                last_change_date_time=NOW,
            )
        )

    asyncio.run(_db_exec(_create))
    return ASSOC_IDENT


# ---------------------------------------------------------------------------
# Environment detection tests
# ---------------------------------------------------------------------------


class TestEnvironmentDetection:
    def test_no_database_url(self, runner, monkeypatch, tmp_path):
        """When neither the env var nor a `.env` file is available, CLI must exit 1."""
        monkeypatch.delenv("DATABASE_URL", raising=False)
        # `_check_db()` falls back to reading `.env` in CWD; run from a tmp dir
        # so the repo's `.env` (used for hybrid dev) is not picked up here.
        monkeypatch.chdir(tmp_path)
        from cli import cli

        result = runner.invoke(cli, ["tenant", "list"])
        assert result.exit_code == 1
        assert "DATABASE_URL" in result.output

    def test_database_url_from_env_file(self, runner, monkeypatch, tmp_path):
        """A `.env` file with DATABASE_URL satisfies `_check_db()` even without the env var."""
        monkeypatch.delenv("DATABASE_URL", raising=False)
        monkeypatch.chdir(tmp_path)
        (tmp_path / ".env").write_text("DATABASE_URL=postgresql+asyncpg://x:y@host/db\n")
        from cli import cli

        result = runner.invoke(cli, ["tenant", "list"])
        # The check passes; the actual DB connect fails (we don't care about that here),
        # so the exit code is non-zero but the "DATABASE_URL is missing" message must not appear.
        assert "DATABASE_URL" not in result.output or "missing" not in result.output.lower()


# ---------------------------------------------------------------------------
# Tenant CRUD tests
# ---------------------------------------------------------------------------


class TestTenantCreate:
    def test_create_public(self, runner, env_docker, clean_db):
        from cli import cli

        result = runner.invoke(cli, ["tenant", "create", "--name", "Test Create"])
        assert result.exit_code == 0
        assert "Created tenant:" in result.output
        assert "Test Create" in result.output
        assert "public" in result.output

    def test_create_private(self, runner, env_docker, clean_db):
        from cli import cli

        result = runner.invoke(cli, ["tenant", "create", "--name", "Private", "--private"])
        assert result.exit_code == 0
        assert "private" in result.output

    def test_create_missing_name(self, runner, env_docker):
        from cli import cli

        result = runner.invoke(cli, ["tenant", "create"])
        assert result.exit_code != 0

    def test_create_with_pinned_id(self, runner, env_docker, clean_db):
        """`tenant create --id <uuid>` lets seed scripts keep a stable tenant URL."""
        from cli import cli

        pinned = "abcd0000-0000-4000-8000-000000000001"
        result = runner.invoke(
            cli,
            ["tenant", "create", "--name", "Pinned", "--id", pinned],
        )
        assert result.exit_code == 0
        assert pinned in result.output

    def test_create_with_pinned_id_duplicate(self, runner, env_docker, clean_db):
        from cli import cli

        pinned = "abcd0000-0000-4000-8000-000000000002"
        first = runner.invoke(
            cli,
            ["tenant", "create", "--name", "First", "--id", pinned],
        )
        assert first.exit_code == 0

        second = runner.invoke(
            cli,
            ["tenant", "create", "--name", "Second", "--id", pinned],
        )
        assert second.exit_code != 0
        # The user-facing error string is what stays stable; translation is in cli_*.json
        assert pinned in (second.stderr + second.output)

    def test_create_with_invalid_id(self, runner, env_docker, clean_db):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "create", "--name", "Bad", "--id", "not-a-uuid"],
        )
        assert result.exit_code != 0

    def test_create_with_slug(self, runner, env_docker, clean_db):
        """`tenant create --slug` stores the slug alongside the UUID."""
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "create", "--name", "Ikenohata University", "--slug", "ikenohata-u"],
        )
        assert result.exit_code == 0
        assert "Ikenohata University" in result.output

        listed = runner.invoke(cli, ["tenant", "list"])
        assert "ikenohata-u" in listed.output

    def test_create_with_duplicate_slug(self, runner, env_docker, clean_db):
        from cli import cli

        first = runner.invoke(
            cli,
            ["tenant", "create", "--name", "First", "--slug", "ikenohata-u"],
        )
        assert first.exit_code == 0

        second = runner.invoke(
            cli,
            ["tenant", "create", "--name", "Second", "--slug", "ikenohata-u"],
        )
        assert second.exit_code != 0
        assert "ikenohata-u" in (second.stderr + second.output)

    @pytest.mark.parametrize(
        "bad_slug",
        [
            "-foo",  # leading hyphen
            "foo-",  # trailing hyphen
            "Foo",  # uppercase
            "a",  # too short
            "health",  # reserved
            "static",  # reserved
            "550e8400-e29b-41d4-a716-446655440000",  # UUID-shaped
        ],
    )
    def test_create_with_invalid_slug(self, runner, env_docker, clean_db, bad_slug: str):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "create", "--name", "x", "--slug", bad_slug],
        )
        assert result.exit_code != 0


class TestTenantList:
    def test_list_empty(self, runner, env_docker, clean_db):
        from cli import cli

        result = runner.invoke(cli, ["tenant", "list"])
        assert result.exit_code == 0
        assert "No tenants found" in result.output

    def test_list_with_tenants(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(cli, ["tenant", "list"])
        assert result.exit_code == 0
        assert "Test Tenant" in result.output

    def test_list_with_docs(self, runner, env_docker, test_document):
        from cli import cli

        result = runner.invoke(cli, ["tenant", "list", "--with-docs"])
        assert result.exit_code == 0
        assert "Test Tenant" in result.output
        assert "Test Document" in result.output


class TestTenantUpdate:
    def test_update_name(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--name", "New Name"],
        )
        assert result.exit_code == 0
        assert "Updated tenant:" in result.output
        assert "New Name" in result.output

    def test_update_private(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--private"],
        )
        assert result.exit_code == 0
        assert "private" in result.output

    def test_update_public(self, runner, env_docker, test_tenant):
        from cli import cli

        # First make private
        runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--private"],
        )
        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--public"],
        )
        assert result.exit_code == 0
        assert "public" in result.output

    def test_update_no_options(self, runner, env_docker):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID)],
        )
        assert result.exit_code == 1
        assert "At least one of" in result.output

    def test_update_private_public_conflict(self, runner, env_docker):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--private", "--public"],
        )
        assert result.exit_code == 1
        assert "mutually exclusive" in result.output

    def test_update_not_found(self, runner, env_docker, clean_db):
        from cli import cli

        fake_uuid = "99999999-9999-9999-9999-999999999999"
        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", fake_uuid, "--name", "X"],
        )
        assert result.exit_code == 1
        assert "Tenant not found" in result.output

    def test_update_invalid_uuid(self, runner, env_docker):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", "bad-uuid", "--name", "X"],
        )
        assert result.exit_code == 1
        assert "Invalid UUID format" in result.output

    def test_update_set_slug(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--slug", "ikenohata-u"],
        )
        assert result.exit_code == 0
        listed = runner.invoke(cli, ["tenant", "list"])
        assert "ikenohata-u" in listed.output

    def test_update_clear_slug(self, runner, env_docker, test_tenant):
        from cli import cli

        # Set a slug, then clear it.
        runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--slug", "to-clear"],
        )
        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--clear-slug"],
        )
        assert result.exit_code == 0
        listed = runner.invoke(cli, ["tenant", "list"])
        assert "to-clear" not in listed.output

    def test_update_slug_and_clear_slug_conflict(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            [
                "tenant",
                "update",
                "--tenant",
                str(TENANT_ID),
                "--slug",
                "x",
                "--clear-slug",
            ],
        )
        assert result.exit_code == 1
        assert "mutually exclusive" in result.output

    def test_update_invalid_slug(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--slug", "Bad-Slug!"],
        )
        assert result.exit_code == 1

    def test_update_display_order(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--display-order", "5"],
        )
        assert result.exit_code == 0
        assert "Updated tenant:" in result.output

    def test_update_clear_order(self, runner, env_docker, test_tenant):
        from cli import cli

        runner.invoke(cli, ["tenant", "update", "--tenant", str(TENANT_ID), "--display-order", "5"])
        result = runner.invoke(cli, ["tenant", "update", "--tenant", str(TENANT_ID), "--clear-order"])
        assert result.exit_code == 0

    def test_update_display_order_clear_conflict(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", str(TENANT_ID), "--display-order", "1", "--clear-order"],
        )
        assert result.exit_code == 1
        assert "mutually exclusive" in result.output


class TestDocUpdate:
    def test_doc_update_display_order(self, runner, env_docker, test_document):
        from cli import cli

        result = runner.invoke(
            cli,
            ["doc", "update", "--tenant", str(TENANT_ID), "--doc", str(DOC_IDENT), "--display-order", "3"],
        )
        assert result.exit_code == 0
        assert "Updated document:" in result.output

    def test_doc_update_clear_order(self, runner, env_docker, test_document):
        from cli import cli

        runner.invoke(
            cli, ["doc", "update", "--tenant", str(TENANT_ID), "--doc", str(DOC_IDENT), "--display-order", "3"]
        )
        result = runner.invoke(
            cli, ["doc", "update", "--tenant", str(TENANT_ID), "--doc", str(DOC_IDENT), "--clear-order"]
        )
        assert result.exit_code == 0

    def test_doc_update_requires_option(self, runner, env_docker, test_document):
        from cli import cli

        result = runner.invoke(cli, ["doc", "update", "--tenant", str(TENANT_ID), "--doc", str(DOC_IDENT)])
        assert result.exit_code == 1
        assert "At least one of" in result.output

    def test_doc_update_conflict(self, runner, env_docker, test_document):
        from cli import cli

        result = runner.invoke(
            cli,
            [
                "doc",
                "update",
                "--tenant",
                str(TENANT_ID),
                "--doc",
                str(DOC_IDENT),
                "--display-order",
                "1",
                "--clear-order",
            ],
        )
        assert result.exit_code == 1
        assert "mutually exclusive" in result.output

    def test_doc_update_not_found(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["doc", "update", "--tenant", str(TENANT_ID), "--doc", str(uuid.uuid4()), "--display-order", "1"],
        )
        assert result.exit_code == 1
        assert "Document not found" in result.output


class TestListDisplayOrder:
    """CLI list commands honor display_order (smaller = higher, NULLs last)."""

    def test_tenant_list_respects_display_order(self, runner, env_docker, clean_db):
        from cli import cli

        apple = "00000000-0000-0000-0000-0000000000a1"
        zebra = "00000000-0000-0000-0000-0000000000a2"
        runner.invoke(cli, ["tenant", "create", "--name", "Apple", "--id", apple])
        runner.invoke(cli, ["tenant", "create", "--name", "Zebra", "--id", zebra])
        # Pin Zebra above the alphabetical default.
        runner.invoke(cli, ["tenant", "update", "--tenant", zebra, "--display-order", "1"])

        result = runner.invoke(cli, ["tenant", "list"])
        assert result.exit_code == 0
        assert result.output.index("Zebra") < result.output.index("Apple")

    def test_doc_list_respects_display_order(self, runner, env_docker, test_tenant):
        from cli import cli

        a_id = uuid.UUID("dddddddd-dddd-dddd-dddd-dddddddddddd")
        z_id = uuid.UUID("eeeeeeee-eeee-eeee-eeee-eeeeeeeeeeee")

        async def _seed(session):
            session.add(
                CFDocument(tenant_id=TENANT_ID, identifier=a_id, uri="ua", title="Apple Doc", last_change_date_time=NOW)
            )
            session.add(
                CFDocument(tenant_id=TENANT_ID, identifier=z_id, uri="uz", title="Zebra Doc", last_change_date_time=NOW)
            )

        asyncio.run(_db_exec(_seed))
        # Pin Zebra Doc above the alphabetical default (also verifies doc update persisted it).
        runner.invoke(cli, ["doc", "update", "--tenant", str(TENANT_ID), "--doc", str(z_id), "--display-order", "1"])

        result = runner.invoke(cli, ["doc", "list", "--tenant", str(TENANT_ID)])
        assert result.exit_code == 0
        assert result.output.index("Zebra Doc") < result.output.index("Apple Doc")


class TestTenantDelete:
    def test_delete_with_force(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "delete", "--tenant", str(TENANT_ID), "--force"],
        )
        assert result.exit_code == 0
        assert "Deleted tenant:" in result.output
        assert "Test Tenant" in result.output

    def test_delete_confirm_yes(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "delete", "--tenant", str(TENANT_ID)],
            input="y\n",
        )
        assert result.exit_code == 0
        assert "Deleted tenant:" in result.output

    def test_delete_confirm_no(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "delete", "--tenant", str(TENANT_ID)],
            input="N\n",
        )
        assert result.exit_code == 2
        assert "Cancelled" in result.output

    def test_delete_not_found(self, runner, env_docker, clean_db):
        from cli import cli

        fake_uuid = "99999999-9999-9999-9999-999999999999"
        result = runner.invoke(
            cli,
            ["tenant", "delete", "--tenant", fake_uuid, "--force"],
        )
        assert result.exit_code == 1
        assert "Tenant not found" in result.output


# ---------------------------------------------------------------------------
# Doc tests
# ---------------------------------------------------------------------------


class TestDocList:
    def test_list_docs(self, runner, env_docker, test_document):
        from cli import cli

        result = runner.invoke(
            cli,
            ["doc", "list", "--tenant", str(TENANT_ID)],
        )
        assert result.exit_code == 0
        assert "Test Document" in result.output

    def test_list_empty(self, runner, env_docker, test_tenant):
        from cli import cli

        result = runner.invoke(
            cli,
            ["doc", "list", "--tenant", str(TENANT_ID)],
        )
        assert result.exit_code == 0
        assert "No documents found" in result.output

    def test_list_tenant_not_found(self, runner, env_docker, clean_db):
        from cli import cli

        fake_uuid = "99999999-9999-9999-9999-999999999999"
        result = runner.invoke(cli, ["doc", "list", "--tenant", fake_uuid])
        assert result.exit_code == 1
        assert "Tenant not found" in result.output


class TestDocDelete:
    def test_delete_with_force(self, runner, env_docker, test_document):
        from cli import cli

        result = runner.invoke(
            cli,
            [
                "doc",
                "delete",
                "--tenant",
                str(TENANT_ID),
                "--doc",
                str(DOC_IDENT),
                "--force",
            ],
        )
        assert result.exit_code == 0
        assert "Deleted document:" in result.output
        assert "Test Document" in result.output

    def test_delete_confirm_cancel(self, runner, env_docker, test_document):
        from cli import cli

        result = runner.invoke(
            cli,
            [
                "doc",
                "delete",
                "--tenant",
                str(TENANT_ID),
                "--doc",
                str(DOC_IDENT),
            ],
            input="N\n",
        )
        assert result.exit_code == 2
        assert "Cancelled" in result.output

    def test_delete_doc_not_found(self, runner, env_docker, test_tenant):
        from cli import cli

        fake_uuid = "99999999-9999-9999-9999-999999999999"
        result = runner.invoke(
            cli,
            [
                "doc",
                "delete",
                "--tenant",
                str(TENANT_ID),
                "--doc",
                fake_uuid,
                "--force",
            ],
        )
        assert result.exit_code == 1
        assert "Document not found" in result.output


# ---------------------------------------------------------------------------
# Import CSV tests
# ---------------------------------------------------------------------------


class TestImportCsv:
    def test_file_not_found(self, runner, env_docker):
        from cli import cli

        result = runner.invoke(
            cli,
            ["import", "csv", "--tenant", str(TENANT_ID), "--file", "/nonexistent.csv"],
        )
        assert result.exit_code == 1
        assert "File not found" in result.output

    def test_invalid_utf8(self, runner, env_docker, tmp_path):
        from cli import cli

        bad_file = tmp_path / "bad.csv"
        bad_file.write_bytes(b"\xff\xfe invalid")
        result = runner.invoke(
            cli,
            ["import", "csv", "--tenant", str(TENANT_ID), "--file", str(bad_file)],
        )
        assert result.exit_code == 1
        assert "CSV file is not valid UTF-8" in result.output

    def test_import_success(self, runner, env_docker, test_tenant, tmp_path):
        from cli import cli

        csv_file = tmp_path / "test.csv"
        csv_file.write_text(
            "#title,Test Framework\n"
            "Identifier,Full Statement,Human Coding Scheme,Parent Identifier\n"
            ",Statement 1,A-1,\n",
            encoding="utf-8",
        )
        result = runner.invoke(
            cli,
            ["import", "csv", "--tenant", str(TENANT_ID), "--file", str(csv_file)],
        )
        assert result.exit_code == 0
        assert "Imported into" in result.output

    def test_import_tenant_not_found(self, runner, env_docker, clean_db, tmp_path):
        from cli import cli

        csv_file = tmp_path / "test.csv"
        csv_file.write_text("Identifier,Full Statement\n", encoding="utf-8")
        fake_uuid = "99999999-9999-9999-9999-999999999999"
        result = runner.invoke(
            cli,
            ["import", "csv", "--tenant", fake_uuid, "--file", str(csv_file)],
        )
        assert result.exit_code == 1
        assert "Tenant not found" in result.output

    def test_import_doc_not_found(self, runner, env_docker, test_tenant, tmp_path):
        from cli import cli

        csv_file = tmp_path / "test.csv"
        csv_file.write_text("Identifier,Full Statement\n", encoding="utf-8")
        fake_uuid = "99999999-9999-9999-9999-999999999999"
        result = runner.invoke(
            cli,
            [
                "import",
                "csv",
                "--tenant",
                str(TENANT_ID),
                "--doc",
                fake_uuid,
                "--file",
                str(csv_file),
            ],
        )
        assert result.exit_code == 1
        assert "Document not found" in result.output

    def test_import_profile_mismatch(self, runner, env_docker, test_tenant, tmp_path):
        # Header detects as OpenSALT ("full statement"), but --profile simple is
        # forced → no silent fallback, clean error exit.
        from cli import cli

        csv_file = tmp_path / "test.csv"
        csv_file.write_text("Identifier,Full Statement\n,Stmt 1\n", encoding="utf-8")
        result = runner.invoke(
            cli,
            [
                "import",
                "csv",
                "--tenant",
                str(TENANT_ID),
                "--file",
                str(csv_file),
                "--profile",
                "simple",
            ],
        )
        assert result.exit_code == 1
        assert "opensalt" in result.output

    def test_import_profile_invalid_value(self, runner, env_docker):
        from cli import cli

        result = runner.invoke(
            cli,
            ["import", "csv", "--tenant", str(TENANT_ID), "--file", "/x.csv", "--profile", "xml"],
        )
        # click.Choice rejects unknown values with a usage error (exit code 2).
        assert result.exit_code == 2


# ---------------------------------------------------------------------------
# Import CASE tests
# ---------------------------------------------------------------------------


class TestImportCase:
    def test_requires_exactly_one_source(self, runner, env_docker):
        from cli import cli

        # Neither --url nor --file.
        result = runner.invoke(cli, ["import", "case", "--tenant", str(TENANT_ID)])
        assert result.exit_code == 1
        assert "exactly one" in result.output

    def test_rejects_both_sources(self, runner, env_docker, tmp_path):
        from cli import cli

        f = tmp_path / "pkg.json"
        f.write_text("{}", encoding="utf-8")
        result = runner.invoke(
            cli,
            ["import", "case", "--tenant", str(TENANT_ID), "--url", "https://x/y", "--file", str(f)],
        )
        assert result.exit_code == 1
        assert "exactly one" in result.output


# ---------------------------------------------------------------------------
# Export CSV tests
# ---------------------------------------------------------------------------


class TestExportCsv:
    def test_export_success(self, runner, env_docker, test_document, tmp_path):
        from cli import cli

        out_file = tmp_path / "out.csv"
        result = runner.invoke(
            cli,
            [
                "export",
                "csv",
                "--tenant",
                str(TENANT_ID),
                "--doc",
                str(DOC_IDENT),
                "--file",
                str(out_file),
            ],
        )
        assert result.exit_code == 0
        assert "Exported" in result.output
        assert out_file.exists()

    def test_export_opensalt_format(self, runner, env_docker, test_document, tmp_path):
        from cli import cli

        out_file = tmp_path / "opensalt.csv"
        result = runner.invoke(
            cli,
            [
                "export",
                "csv",
                "--tenant",
                str(TENANT_ID),
                "--doc",
                str(DOC_IDENT),
                "--file",
                str(out_file),
                "--profile",
                "opensalt",
            ],
        )
        assert result.exit_code == 0
        assert "Exported" in result.output
        assert out_file.exists()
        content = out_file.read_text()
        assert "Is Child Of" in content

    def test_export_invalid_profile(self, runner, env_docker):
        from cli import cli

        result = runner.invoke(
            cli,
            [
                "export",
                "csv",
                "--tenant",
                str(TENANT_ID),
                "--doc",
                str(DOC_IDENT),
                "--file",
                "/tmp/out.csv",
                "--profile",
                "xml",
            ],
        )
        assert result.exit_code == 1
        assert "Invalid profile: 'xml'" in result.output

    def test_export_tenant_not_found(self, runner, env_docker, clean_db, tmp_path):
        from cli import cli

        fake_uuid = "99999999-9999-9999-9999-999999999999"
        result = runner.invoke(
            cli,
            [
                "export",
                "csv",
                "--tenant",
                fake_uuid,
                "--doc",
                str(DOC_IDENT),
                "--file",
                str(tmp_path / "out.csv"),
            ],
        )
        assert result.exit_code == 1
        assert "Tenant not found" in result.output

    def test_export_doc_not_found(self, runner, env_docker, test_tenant, tmp_path):
        from cli import cli

        fake_uuid = "99999999-9999-9999-9999-999999999999"
        result = runner.invoke(
            cli,
            [
                "export",
                "csv",
                "--tenant",
                str(TENANT_ID),
                "--doc",
                fake_uuid,
                "--file",
                str(tmp_path / "out.csv"),
            ],
        )
        assert result.exit_code == 1
        assert "Document not found" in result.output


# ---------------------------------------------------------------------------
# DB migrate
# ---------------------------------------------------------------------------


class TestDbMigrate:
    def test_migrate_runs_alembic(self, runner, env_docker):
        from cli import cli

        result = runner.invoke(cli, ["db", "migrate"])
        assert result.exit_code == 0
        assert "Migration complete" in result.output


# ---------------------------------------------------------------------------
# Cache invalidate
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# UUID validation
# ---------------------------------------------------------------------------


class TestUuidValidation:
    def test_invalid_tenant_uuid_on_update(self, runner, env_docker):
        from cli import cli

        result = runner.invoke(
            cli,
            ["tenant", "update", "--tenant", "not-a-uuid", "--name", "X"],
        )
        assert result.exit_code == 1
        assert "Invalid UUID format" in result.output

    def test_invalid_doc_uuid_on_delete(self, runner, env_docker):
        from cli import cli

        result = runner.invoke(
            cli,
            [
                "doc",
                "delete",
                "--tenant",
                str(TENANT_ID),
                "--doc",
                "bad-uuid",
                "--force",
            ],
        )
        assert result.exit_code == 1
        assert "Invalid UUID format" in result.output


# ---------------------------------------------------------------------------
# XLSX import / export tests
# ---------------------------------------------------------------------------


class TestXlsxCli:
    def test_export_xlsx_success(self, runner, env_docker, test_document, tmp_path):
        from cli import cli

        out_file = tmp_path / "out.xlsx"
        result = runner.invoke(
            cli,
            ["export", "xlsx", "--tenant", str(TENANT_ID), "--doc", str(DOC_IDENT), "--file", str(out_file)],
        )
        assert result.exit_code == 0
        assert out_file.exists()
        from openpyxl import load_workbook

        wb = load_workbook(out_file)
        assert wb.sheetnames == ["CF Doc", "CF Item", "CF Association"]

    def test_import_xlsx_success(self, runner, env_docker, test_tenant, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        wb.remove(wb.active)
        d = wb.create_sheet("CF Doc")
        d.append(["identifier", "creator", "title"] + [""] * 13)
        d.append([str(uuid.uuid4()), "Author", "Imported via CLI"] + [""] * 13)
        it = wb.create_sheet("CF Item")
        it.append(["identifier", "fullStatement", "humanCodingScheme", "smartLevel"] + [""] * 8)
        it.append([str(uuid.uuid4()), "Root item", "R", "1"] + [""] * 8)
        wb.create_sheet("CF Association").append(["identifier"] + [""] * 9)
        path = tmp_path / "in.xlsx"
        wb.save(path)

        from cli import cli

        result = runner.invoke(cli, ["import", "xlsx", "--tenant", str(TENANT_ID), "--file", str(path)])
        assert result.exit_code == 0, result.output
        assert "Imported into" in result.output

    def test_import_xlsx_bad_file(self, runner, env_docker, test_tenant, tmp_path):
        bad = tmp_path / "bad.xlsx"
        bad.write_bytes(b"not a real xlsx")
        from cli import cli

        result = runner.invoke(cli, ["import", "xlsx", "--tenant", str(TENANT_ID), "--file", str(bad)])
        assert result.exit_code == 1


def _assoc_count() -> int:
    """How many rows the association identifier still has (0 or 1)."""
    found = []

    async def _count(session):
        from sqlalchemy import func

        result = await session.execute(
            select(func.count(CFAssociation.id)).where(CFAssociation.identifier == ASSOC_IDENT)
        )
        found.append(result.scalar())

    asyncio.run(_db_exec(_count))
    return found[0]


class TestAssocDelete:
    """B8-7: taking back a wrong `replacedBy`.

    Import is additive for associations — a package cannot say "this link is
    gone" — so a wrong successor link survives every re-import and keeps sending
    readers to the wrong item. This command is the way back.
    """

    def _invoke(self, runner, ident, *extra, input=None):
        from cli import cli

        return runner.invoke(
            cli, ["assoc", "delete", "--tenant", str(TENANT_ID), "--id", str(ident), *extra], input=input
        )

    def test_delete_with_force(self, runner, env_docker, test_association):
        result = self._invoke(runner, test_association, "--force")
        assert result.exit_code == 0
        assert "replacedBy" in result.output
        assert _assoc_count() == 0

    def test_confirm_shows_both_endpoints(self, runner, env_docker, test_association):
        """The operator has to see what the link points at before removing it."""
        result = self._invoke(runner, test_association, input="N\n")
        assert result.exit_code == 2
        assert "11111111-2222-3333-4444-555555555555" in result.output
        assert "66666666-7777-8888-9999-000000000000" in result.output

    def test_cancel_keeps_it(self, runner, env_docker, test_association):
        self._invoke(runner, test_association, input="N\n")
        assert _assoc_count() == 1

    def test_not_found(self, runner, env_docker, test_tenant):
        result = self._invoke(runner, uuid.uuid4(), "--force")
        assert result.exit_code == 1

    def test_other_tenant_cannot_delete_it(self, runner, env_docker, test_association):
        """Tenant scope is enforced, not just tenant existence.

        A nonexistent tenant would exit at the existence check without ever
        reaching the association query — that passes even with the scope removed.
        The second tenant here has to be real.
        """
        from cli import cli

        other_tenant = uuid.uuid4()

        async def _create(session):
            session.add(Tenant(id=other_tenant, name="Other", is_private=False))

        asyncio.run(_db_exec(_create))

        result = runner.invoke(
            cli,
            ["assoc", "delete", "--tenant", str(other_tenant), "--id", str(test_association), "--force"],
        )
        assert result.exit_code == 1
        assert _assoc_count() == 1

    def test_is_child_of_is_refused(self, runner, env_docker, test_document):
        """Deleting a hierarchy link would drop the item out of the tree.

        `cf_items.depth` is stored and only recomputed on import, and the tree
        finds roots and orphans by `depth == 0`; a cut child sits at depth >= 1
        with no parent and disappears from the Web UI until the next import.
        """
        ident = uuid.uuid4()

        async def _create(session):
            doc = (await session.execute(select(CFDocument).where(CFDocument.identifier == DOC_IDENT))).scalar_one()
            session.add(
                CFAssociation(
                    tenant_id=TENANT_ID,
                    cf_document_id=doc.id,
                    identifier=ident,
                    uri=f"https://example.com/assoc/{ident}",
                    association_type="isChildOf",
                    origin_node_uri="https://example.com/uri/child",
                    origin_node_identifier=str(uuid.uuid4()),
                    destination_node_uri="https://example.com/uri/parent",
                    destination_node_identifier=str(uuid.uuid4()),
                    last_change_date_time=NOW,
                )
            )

        asyncio.run(_db_exec(_create))

        from cli import cli

        result = runner.invoke(cli, ["assoc", "delete", "--tenant", str(TENANT_ID), "--id", str(ident), "--force"])
        assert result.exit_code == 1

        found = []

        async def _count(session):
            from sqlalchemy import func

            found.append(
                (
                    await session.execute(select(func.count(CFAssociation.id)).where(CFAssociation.identifier == ident))
                ).scalar()
            )

        asyncio.run(_db_exec(_count))
        assert found == [1]

    def test_prompt_shows_endpoint_titles(self, runner, env_docker, test_association):
        """Two bare UUIDs are not enough to decide whether a link should go."""
        result = self._invoke(runner, test_association, input="N\n")
        assert "Old item" in result.output
        assert "New item" in result.output
        assert "Test Document" in result.output
