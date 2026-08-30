"""Unit tests for the XLSX import/export services (OpenSALT Excel format).

The core check is a round-trip: build a framework via the custom CSV importer,
export it to an OpenSALT-format .xlsx, re-import that workbook into a fresh
document, and assert the items / hierarchy / item types / education levels and
the non-isChildOf associations survive.
"""

from __future__ import annotations

import io
import uuid
from datetime import date

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from src.models.cf_association import CFAssociation
from src.models.cf_item import CFItem
from src.models.tenant import Tenant
from src.services.csv_import_service import import_csv
from src.services.xlsx_export_service import export_xlsx
from src.services.xlsx_import_service import import_xlsx

TENANT_ID = uuid.UUID("11111111-1111-1111-1111-111111111111")

# A small custom-format framework: 2 roots, one with a child; item types and
# education levels set; plus one non-isChildOf (isRelatedTo) association.
SOURCE_CSV = (
    "#identifier,dddddddd-0000-0000-0000-000000000001\n"
    "#title,XLSX Source Framework\n"
    "#creator,Test Author\n"
    "#language,ja\n"
    "#version,1.0\n"
    "#subject,情報科学,データ\n"
    "Identifier,fullStatement,humanCodingScheme,parentIdentifier,sequenceNumber,CFItemType,educationLevel,conceptKeywords,abbreviatedStatement,language,listEnumeration,license,statusStartDate,statusEndDate\n"  # noqa: E501
    "10000000-0000-0000-0000-000000000001,Root A,A,,10,領域,13,,Root A short,,,,,\n"
    '10000000-0000-0000-0000-000000000002,Child A1,A-1,10000000-0000-0000-0000-000000000001,10,知識,"13,14","kw1,kw2",Child A1 short,,,,2021-04-01,2022-03-14\n'  # noqa: E501
    "10000000-0000-0000-0000-000000000003,Root B,B,,20,領域,14,,Root B short,,,,,\n"
)


async def _seed_source(session: AsyncSession) -> uuid.UUID:
    session.add(Tenant(id=TENANT_ID, name="T", is_private=False))
    await session.flush()
    report = await import_csv(session, TENANT_ID, SOURCE_CSV.encode("utf-8"))
    await session.flush()
    # Add a non-isChildOf association: Child A1 isRelatedTo Root B.
    doc_id = uuid.UUID(report.document_identifier)
    from src.models.cf_document import CFDocument

    doc = (await session.execute(select(CFDocument).where(CFDocument.identifier == doc_id))).scalar_one()
    a1 = uuid.UUID("10000000-0000-0000-0000-000000000002")
    b = uuid.UUID("10000000-0000-0000-0000-000000000003")
    session.add(
        CFAssociation(
            id=uuid.uuid4(),
            tenant_id=TENANT_ID,
            cf_document_id=doc.id,
            identifier=uuid.UUID("20000000-0000-0000-0000-0000000000aa"),
            uri="https://example.com/assoc/aa",
            association_type="isRelatedTo",
            origin_node_uri=f"https://example.com/{a1}",
            origin_node_identifier=str(a1),
            destination_node_uri=f"https://example.com/{b}",
            destination_node_identifier=str(b),
            last_change_date_time=doc.last_change_date_time,
        )
    )
    await session.flush()
    return doc_id


class TestXlsxExport:
    async def test_export_workbook_structure(self, db_session: AsyncSession):
        doc_id = await _seed_source(db_session)
        data = await export_xlsx(db_session, TENANT_ID, doc_id)

        wb = load_workbook(io.BytesIO(data))
        assert wb.sheetnames == ["CF Doc", "CF Item", "CF Association"]

        # CF Doc: header + 1 data row
        doc_ws = wb["CF Doc"]
        assert doc_ws.cell(1, 1).value == "identifier"
        assert doc_ws.cell(2, 3).value == "XLSX Source Framework"  # title col C
        assert doc_ws.cell(2, 8).value == "情報科学|データ"  # subject pipe-joined

        # CF Item: 3 data rows with smartLevel
        item_ws = wb["CF Item"]
        rows = list(item_ws.iter_rows(min_row=2, values_only=True))
        by_stmt = {r[1]: r for r in rows}
        assert by_stmt["Root A"][3] == "1"  # smartLevel col D
        assert by_stmt["Child A1"][3] == "1.1"
        assert by_stmt["Root B"][3] == "2"
        assert by_stmt["Child A1"][10] == "知識"  # CFItemType col K
        assert by_stmt["Child A1"][9] == "13,14"  # educationLevel col J
        # Lifecycle dates (compeito extension, cols M-N): without them the
        # retirement state is lost on an export -> re-import round trip.
        assert by_stmt["Child A1"][12] == "2021-04-01"
        assert by_stmt["Child A1"][13] == "2022-03-14"
        assert not by_stmt["Root A"][13]  # openpyxl reads an empty cell as None

        # CF Association: the isRelatedTo (isChildOf NOT repeated here)
        assoc_ws = wb["CF Association"]
        arows = list(assoc_ws.iter_rows(min_row=2, values_only=True))
        assert len(arows) == 1
        assert arows[0][4] == "isRelatedTo"  # associationType col E

    async def test_export_doc_not_found(self, db_session: AsyncSession):
        db_session.add(Tenant(id=TENANT_ID, name="T", is_private=False))
        await db_session.flush()
        with pytest.raises(ValueError):
            await export_xlsx(db_session, TENANT_ID, uuid.uuid4())


class TestXlsxRoundTrip:
    async def test_roundtrip_into_new_document(self, db_session: AsyncSession):
        src_doc = await _seed_source(db_session)
        data = await export_xlsx(db_session, TENANT_ID, src_doc)

        # Wipe the stored dates so the re-imported values can only come from the
        # workbook (an empty cell would preserve, hiding a missing column).
        child_before = (
            await db_session.execute(
                select(CFItem).where(CFItem.identifier == uuid.UUID("10000000-0000-0000-0000-000000000002"))
            )
        ).scalar_one()
        child_before.status_start_date = None
        child_before.status_end_date = None
        await db_session.flush()

        # Re-import the workbook (same identifiers → upsert in place).
        report = await import_xlsx(db_session, TENANT_ID, data)
        await db_session.flush()
        assert report.document_identifier == str(src_doc)

        from src.models.cf_document import CFDocument

        doc = (await db_session.execute(select(CFDocument).where(CFDocument.identifier == src_doc))).scalar_one()

        items = list(
            (
                await db_session.execute(
                    select(CFItem).options(joinedload(CFItem.item_type)).where(CFItem.cf_document_id == doc.id)
                )
            )
            .scalars()
            .unique()
        )
        assert len(items) == 3
        by_stmt = {i.full_statement: i for i in items}
        # Item type + education level survived the round-trip.
        child = by_stmt["Child A1"]
        assert child.item_type is not None and child.item_type.title == "知識"
        assert child.education_level == ["13", "14"]
        # Retirement state survives the round-trip (B8-2): without cols M-N the
        # tombstone would come back as a live item.
        assert child.status_start_date == date(2021, 4, 1)
        assert child.status_end_date == date(2022, 3, 14)

        # isChildOf rebuilt from smartLevel: Child A1 → Root A.
        ischild = list(
            (
                await db_session.execute(
                    select(CFAssociation).where(
                        CFAssociation.cf_document_id == doc.id,
                        CFAssociation.association_type == "isChildOf",
                    )
                )
            ).scalars()
        )
        root_a = by_stmt["Root A"]
        child_links = [a for a in ischild if a.origin_node_identifier == str(child.identifier)]
        assert len(child_links) == 1
        assert child_links[0].destination_node_identifier == str(root_a.identifier)

        # Non-isChildOf association preserved.
        related = list(
            (
                await db_session.execute(
                    select(CFAssociation).where(
                        CFAssociation.cf_document_id == doc.id,
                        CFAssociation.association_type == "isRelatedTo",
                    )
                )
            ).scalars()
        )
        assert len(related) == 1


class TestXlsxOpenSaltCompatibility:
    async def test_workbook_without_lifecycle_columns_preserves_dates(self, db_session: AsyncSession):
        """A genuine OpenSALT workbook stops at col L and must not wipe the dates.

        Missing columns are padded to "" by the reader, which the CSV path
        treats as "no value → preserve".
        """
        src_doc = await _seed_source(db_session)
        data = await export_xlsx(db_session, TENANT_ID, src_doc)

        # Truncate the CF Item sheet back to the 12-column OpenSALT layout.
        wb = load_workbook(io.BytesIO(data))
        ws = wb["CF Item"]
        ws.delete_cols(13, 2)
        buf = io.BytesIO()
        wb.save(buf)

        await import_xlsx(db_session, TENANT_ID, buf.getvalue())
        await db_session.flush()

        child = (
            await db_session.execute(
                select(CFItem).where(CFItem.identifier == uuid.UUID("10000000-0000-0000-0000-000000000002"))
            )
        ).scalar_one()
        assert child.status_start_date == date(2021, 4, 1)
        assert child.status_end_date == date(2022, 3, 14)


class TestXlsxLifecycleColumns:
    """The compeito extension columns (statusStartDate / statusEndDate)."""

    async def _export_with_item_sheet(self, db_session: AsyncSession, mutate):
        src_doc = await _seed_source(db_session)
        data = await export_xlsx(db_session, TENANT_ID, src_doc)
        wb = load_workbook(io.BytesIO(data))
        mutate(wb["CF Item"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def _child(self, db_session: AsyncSession) -> CFItem:
        return (
            await db_session.execute(
                select(CFItem).where(CFItem.identifier == uuid.UUID("10000000-0000-0000-0000-000000000002"))
            )
        ).scalar_one()

    async def test_columns_are_found_by_header_name(self, db_session: AsyncSession):
        """OpenSALT emits its AdditionalFields from col 13 in arbitrary order.

        A fixed index would read someone else's custom field as a date, so the
        columns must be located by header name even when they move.
        """

        def move_columns(ws):
            ws.insert_cols(13)
            ws.cell(1, 13).value = "difficulty"
            for row in range(2, ws.max_row + 1):
                ws.cell(row, 13).value = "high"

        data = await self._export_with_item_sheet(db_session, move_columns)
        child = await self._child(db_session)
        child.status_start_date = None
        child.status_end_date = None
        await db_session.flush()

        await import_xlsx(db_session, TENANT_ID, data)
        await db_session.flush()
        await db_session.refresh(child)
        assert child.status_start_date == date(2021, 4, 1)
        assert child.status_end_date == date(2022, 3, 14)

    async def test_unregistered_custom_field_is_not_read_as_a_date(self, db_session: AsyncSession):
        """A workbook with a custom field but no lifecycle headers preserves."""

        def replace_headers(ws):
            ws.cell(1, 13).value = "difficulty"
            ws.cell(1, 14).value = "reviewedBy"

        data = await self._export_with_item_sheet(db_session, replace_headers)
        await import_xlsx(db_session, TENANT_ID, data)
        await db_session.flush()

        child = await self._child(db_session)
        assert child.status_start_date == date(2021, 4, 1)
        assert child.status_end_date == date(2022, 3, 14)

    async def test_date_formatted_cells(self, db_session: AsyncSession):
        """Editing the workbook in Excel turns the strings into date cells."""

        def to_date_cells(ws):
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 13).value:
                    ws.cell(row, 13).value = date(2021, 4, 1)
                if ws.cell(row, 14).value:
                    ws.cell(row, 14).value = date(2022, 3, 14)

        data = await self._export_with_item_sheet(db_session, to_date_cells)
        child = await self._child(db_session)
        child.status_start_date = None
        child.status_end_date = None
        await db_session.flush()

        report = await import_xlsx(db_session, TENANT_ID, data)
        await db_session.flush()
        await db_session.refresh(child)
        assert child.status_start_date == date(2021, 4, 1)
        assert child.status_end_date == date(2022, 3, 14)
        assert not any("Invalid status" in w for w in report.warnings)

    async def test_duplicate_header_warns_and_uses_the_leftmost(self, db_session: AsyncSession):
        """Two statusEndDate columns are ambiguous — warn instead of guessing."""

        def duplicate(ws):
            ws.cell(1, 15).value = "statusEndDate"
            for row in range(2, ws.max_row + 1):
                ws.cell(row, 15).value = "2099-12-31"

        data = await self._export_with_item_sheet(db_session, duplicate)
        report = await import_xlsx(db_session, TENANT_ID, data)
        await db_session.flush()

        child = await self._child(db_session)
        assert child.status_end_date == date(2022, 3, 14)  # leftmost column
        assert any("duplicate" in w for w in report.warnings)

    async def test_invalid_value_keeps_the_stored_date(self, db_session: AsyncSession):
        """A malformed cell warns; it must not wipe the retirement date.

        The xlsx / CSV path has no explicit clearing mechanism, so "n/a" cannot
        be an authoritative "remove the retirement date" instruction.
        """

        def break_cells(ws):
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, 14).value:
                    ws.cell(row, 14).value = "n/a"

        data = await self._export_with_item_sheet(db_session, break_cells)
        report = await import_xlsx(db_session, TENANT_ID, data)
        await db_session.flush()

        child = await self._child(db_session)
        assert child.status_end_date == date(2022, 3, 14)
        assert any("Invalid statusEndDate" in w and "kept" in w for w in report.warnings)


class TestXlsxDocumentLifecycleDates:
    """CF Doc sheet lifecycle dates (the #status_* metadata rows)."""

    async def _doc(self, db_session: AsyncSession):
        from src.models.cf_document import CFDocument

        return (
            await db_session.execute(
                select(CFDocument).where(CFDocument.identifier == uuid.UUID("dddddddd-0000-0000-0000-000000000001"))
            )
        ).scalar_one()

    async def test_invalid_doc_date_keeps_stored_value(self, db_session: AsyncSession):
        """The #identifier branch must not wipe the document's retirement date.

        A plain `import xlsx` (no --doc) resolves the document from the CF Doc
        sheet's identifier, which is a different code path from --doc.
        """
        src_doc = await _seed_source(db_session)
        doc = await self._doc(db_session)
        doc.status_end_date = date(2026, 3, 31)
        await db_session.flush()

        data = await export_xlsx(db_session, TENANT_ID, src_doc)
        wb = load_workbook(io.BytesIO(data))
        wb["CF Doc"].cell(2, 13).value = "n/a"  # statusEndDate (CF Doc col M)
        buf = io.BytesIO()
        wb.save(buf)

        report = await import_xlsx(db_session, TENANT_ID, buf.getvalue())
        await db_session.flush()
        await db_session.refresh(doc)
        assert doc.status_end_date == date(2026, 3, 31)
        assert any("#status_end_date" in w for w in report.warnings)

    async def test_invalid_doc_date_keeps_stored_value_with_doc_flag(self, db_session: AsyncSession):
        """Same rule on the --doc branch."""
        src_doc = await _seed_source(db_session)
        doc = await self._doc(db_session)
        doc.status_end_date = date(2026, 3, 31)
        await db_session.flush()

        data = await export_xlsx(db_session, TENANT_ID, src_doc)
        wb = load_workbook(io.BytesIO(data))
        wb["CF Doc"].cell(2, 13).value = "n/a"
        buf = io.BytesIO()
        wb.save(buf)

        await import_xlsx(db_session, TENANT_ID, buf.getvalue(), doc_identifier=src_doc)
        await db_session.flush()
        await db_session.refresh(doc)
        assert doc.status_end_date == date(2026, 3, 31)


class TestXlsxAssociationGrouping:
    async def test_import_association_with_grouping(self, db_session: AsyncSession):
        """A CF Association row carrying an associationGroupIdentifier must
        find-or-create a (tenant-wide) CFAssociationGrouping — regression guard
        for the document-scoped lookup bug."""
        from openpyxl import Workbook

        from src.models.cf_association_grouping import CFAssociationGrouping

        db_session.add(Tenant(id=TENANT_ID, name="T", is_private=False))
        await db_session.flush()

        origin = uuid.UUID("30000000-0000-0000-0000-000000000001")
        dest = uuid.UUID("30000000-0000-0000-0000-000000000002")
        group = uuid.UUID("40000000-0000-0000-0000-0000000000aa")

        wb = Workbook()
        wb.remove(wb.active)
        d = wb.create_sheet("CF Doc")
        d.append(["identifier", "creator", "title"] + [""] * 13)
        d.append([str(uuid.uuid4()), "A", "Grouping Test"] + [""] * 13)
        it = wb.create_sheet("CF Item")
        it.append(["identifier", "fullStatement", "humanCodingScheme", "smartLevel"] + [""] * 8)
        it.append([str(origin), "Origin item", "O", "1"] + [""] * 8)
        it.append([str(dest), "Dest item", "D", "2"] + [""] * 8)
        a = wb.create_sheet("CF Association")
        a.append(
            [
                "identifier",
                "originNodeURI",
                "originNodeIdentifier",
                "originNodeHumanCodingScheme",
                "associationType",
                "destinationNodeURI",
                "destinationNodeIdentifier",
                "destinationNodeHumanCodingScheme",
                "associationGroupIdentifier",
                "associationGroupName",
            ]
        )
        a.append(
            [
                str(uuid.uuid4()),
                "",
                str(origin),
                "",
                "isRelatedTo",
                "",
                str(dest),
                "",
                str(group),
                "Crosswalk Group",
            ]
        )
        buf = io.BytesIO()
        wb.save(buf)

        await import_xlsx(db_session, TENANT_ID, buf.getvalue())
        await db_session.flush()

        groupings = list(
            (
                await db_session.execute(
                    select(CFAssociationGrouping).where(CFAssociationGrouping.tenant_id == TENANT_ID)
                )
            ).scalars()
        )
        assert len(groupings) == 1
        assert groupings[0].identifier == group
        assert groupings[0].title == "Crosswalk Group"

        rel = (
            await db_session.execute(select(CFAssociation).where(CFAssociation.association_type == "isRelatedTo"))
        ).scalar_one()
        assert rel.cf_association_grouping_id == groupings[0].id


class TestXlsxImportErrors:
    async def test_non_xlsx_bytes(self, db_session: AsyncSession):
        db_session.add(Tenant(id=TENANT_ID, name="T", is_private=False))
        await db_session.flush()
        with pytest.raises(ValueError):
            await import_xlsx(db_session, TENANT_ID, b"not an xlsx file")

    async def test_missing_required_sheet(self, db_session: AsyncSession):
        db_session.add(Tenant(id=TENANT_ID, name="T", is_private=False))
        await db_session.flush()
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.title = "CF Doc"  # missing CF Item
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(ValueError, match="CF Item"):
            await import_xlsx(db_session, TENANT_ID, buf.getvalue())


class TestXlsxDestructiveGuard:
    """An untouched xlsx round trip must not look destructive (B6).

    The generated custom CSV carries every association column, so Step 7.5
    deletes all of them and the CSV stage re-creates none; the second pass
    rebuilds them from the CF Association sheet afterwards. Counting the loss
    before that pass reported every association as lost — exactly the "fires on
    every import" failure the guard was designed to avoid.
    """

    async def test_round_trip_reports_no_loss(self, db_session: AsyncSession):
        src_doc = await _seed_source(db_session)
        data = await export_xlsx(db_session, TENANT_ID, src_doc)

        report = await import_xlsx(db_session, TENANT_ID, data)
        await db_session.flush()

        assert report.lost_associations_count == 0, report.lost_associations_sample
        assert not any("deleted and not re-created" in w for w in report.warnings)

    async def test_a_real_loss_is_still_reported(self, db_session: AsyncSession):
        """The counter still works on this path; it is not simply disabled."""
        src_doc = await _seed_source(db_session)
        data = await export_xlsx(db_session, TENANT_ID, src_doc)

        # Drop the CF Association sheet's only data row: that link is now gone
        # from both passes, so it is a genuine loss.
        wb = load_workbook(io.BytesIO(data))
        wb["CF Association"].delete_rows(2)
        buf = io.BytesIO()
        wb.save(buf)

        report = await import_xlsx(db_session, TENANT_ID, buf.getvalue())
        await db_session.flush()

        assert report.lost_associations_count == 1
        assert report.lost_associations_sample[0][0] == "isRelatedTo"
